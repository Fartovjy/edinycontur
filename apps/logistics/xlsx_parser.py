from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook


@dataclass(frozen=True)
class ParsedCargoItem:
    name: str
    qty: str = ""
    weight_kg: Decimal = Decimal("0")
    volume_m3: Decimal = Decimal("0")
    places_count: int = 0
    needs_supply: bool = True
    needs_cz: bool = False


def _clean(value):
    if value is None:
        return ""
    return str(value).strip()


def _decimal(value):
    if value is None or value == "":
        return Decimal("0")
    try:
        return Decimal(str(value).replace(",", "."))
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _int(value):
    if value is None or value == "":
        return 0
    try:
        return int(Decimal(str(value).replace(",", ".")))
    except (InvalidOperation, ValueError):
        return 0


def _bool_yes(value):
    return _clean(value).lower() in {"да", "yes", "true", "1", "+"}


def _date(value):
    if isinstance(value, datetime):
        return value.date()
    text = _clean(value)
    if not text:
        return None
    for fmt in ("%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def parse_order_xlsx(file_obj):
    workbook = load_workbook(file_obj, data_only=True, read_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    fields = {}
    items = []
    in_items = False

    for row in worksheet.iter_rows(values_only=True):
        first = _clean(row[0] if len(row) > 0 else "")
        second = row[1] if len(row) > 1 else ""
        if not first:
            continue

        if first.lower() == "перечень товаров":
            in_items = True
            continue

        if in_items and first.isdigit():
            name = _clean(second)
            if not name:
                continue
            qty = _clean(row[2] if len(row) > 2 else "")
            item = ParsedCargoItem(
                name=name,
                qty=qty,
                weight_kg=_decimal(row[3] if len(row) > 3 else ""),
                volume_m3=_decimal(row[4] if len(row) > 4 else ""),
                places_count=_int(row[5] if len(row) > 5 else ""),
                needs_supply=_bool_yes(row[6] if len(row) > 6 else ""),
                needs_cz=_bool_yes(row[7] if len(row) > 7 else ""),
            )
            items.append(item)
            continue

        if in_items and first.lower() in {"примечание", "приоритет"}:
            fields[first.lower()] = second
            continue

        if not in_items:
            fields[first.lower()] = second

    total_weight = sum((item.weight_kg for item in items), Decimal("0"))
    total_volume = sum((item.volume_m3 for item in items), Decimal("0"))
    total_places = sum((item.places_count for item in items), 0) or 1
    priority_raw = _clean(fields.get("приоритет")).lower()
    priority = "urgent" if priority_raw in {"высокий", "срочный"} else "normal"

    return {
        "order_number": _clean(fields.get("номер")),
        "order_date": _date(fields.get("дата")),
        "planned_delivery_date": _date(fields.get("плановая доставка")),
        "client_name": _clean(fields.get("клиент")),
        "client_address": _clean(fields.get("адрес / gps")),
        "client_contact": _clean(fields.get("контактное лицо")),
        "client_phone": _clean(fields.get("телефон / email")),
        "note": _clean(fields.get("примечание")),
        "priority": priority,
        "cargo_places_count": total_places,
        "cargo_weight_kg": total_weight,
        "cargo_volume_m3": total_volume,
        "cz_required": any(item.needs_cz for item in items),
        "items": items,
    }
